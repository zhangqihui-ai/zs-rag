from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest

from app.core.document_parser import (
    _is_shared_strings_archive_error,
    _parse_excel_html_table,
    _parse_xlsx_workbook_sheets,
    _prepare_xlsx_zip_bytes,
    _xlsx_zip_has_backslash_paths,
)


def test_is_shared_strings_archive_error():
    assert _is_shared_strings_archive_error(
        KeyError("There is no item named 'xl/sharedStrings.xml' in the archive")
    )
    assert not _is_shared_strings_archive_error(KeyError("other"))


def test_parse_xlsx_workbook_sheets_falls_back_when_shared_strings_missing():
    sheet = MagicMock()
    sheet.title = "Sheet1"
    sheet.iter_rows.return_value = [[("a",), ("b",)]]

    wb_readonly = MagicMock()
    wb_readonly.__iter__.return_value = iter([sheet])

    wb_standard = MagicMock()
    wb_standard.__iter__.return_value = iter([sheet])

    calls: list[bool] = []
    collected: list[str] = []

    def fake_load(_stream, *, read_only, data_only):
        calls.append(read_only)
        if read_only:
            raise KeyError("There is no item named 'xl/sharedStrings.xml' in the archive")
        return wb_standard

    with patch("app.core.document_parser.load_workbook", side_effect=fake_load):
        _parse_xlsx_workbook_sheets(
            b"fake-xlsx",
            "xlsx",
            None,
            engine_label="HTML 表格",
            on_sheet=lambda title, rows: collected.append(title),
        )

    assert calls == [True, False]
    assert collected == ["Sheet1"]
    wb_standard.close.assert_called_once()


def test_prepare_xlsx_zip_bytes_normalizes_backslash_paths():
    import zipfile

    source = BytesIO()
    with zipfile.ZipFile(source, "w") as zout:
        zout.writestr("xl\\sharedStrings.xml", "<sst/>")
        zout.writestr("xl\\worksheets\\sheet1.xml", "<worksheet/>")
    raw = source.getvalue()

    assert _xlsx_zip_has_backslash_paths(raw)
    fixed = _prepare_xlsx_zip_bytes(raw)
    with zipfile.ZipFile(BytesIO(fixed), "r") as zin:
        assert "xl/sharedStrings.xml" in zin.namelist()


def test_parse_excel_html_table_real_backslash_xlsx():
    import zipfile

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "SheetA"
    ws["A1"] = "hello"
    good = BytesIO()
    wb.save(good)
    good_bytes = good.getvalue()

    # Re-pack with Windows-style paths inside zip
    bad = BytesIO()
    with zipfile.ZipFile(BytesIO(good_bytes), "r") as zin, zipfile.ZipFile(bad, "w") as zout:
        for item in zin.infolist():
            zout.writestr(item.filename.replace("/", "\\"), zin.read(item.filename))
    bad_bytes = bad.getvalue()

    logs: list[str] = []
    parsed = _parse_excel_html_table(bad_bytes, "xlsx", log=logs.append)
    assert parsed.segments
    assert "hello" in parsed.text
    assert any("反斜杠" in line for line in logs)


def test_parse_excel_html_table_with_shared_strings_fallback():
    sheet = MagicMock()
    sheet.title = "数据"
    sheet.iter_rows.return_value = [
        ("列1", "列2"),
        ("值1", "值2"),
    ]

    wb_standard = MagicMock()
    wb_standard.__iter__.return_value = iter([sheet])

    def fake_load(_stream, *, read_only, data_only):
        if read_only:
            raise KeyError("There is no item named 'xl/sharedStrings.xml' in the archive")
        return wb_standard

    with patch("app.core.document_parser.load_workbook", side_effect=fake_load):
        parsed = _parse_excel_html_table(b"fake-xlsx", "xlsx")

    assert parsed.parser_type == "xlsx"
    assert parsed.segments
    assert "值1" in parsed.text
